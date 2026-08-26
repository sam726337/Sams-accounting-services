from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Callable


HEADER_ALIASES = {
    "supplier_gstin": (
        "gstin of supplier",
        "supplier gstin",
        "gstin",
        "ctin",
        "gstin/uin of supplier",
    ),
    "supplier_name": (
        "trade/legal name",
        "trade name",
        "legal name",
        "supplier name",
        "name of supplier",
        "party name",
    ),
    "invoice_number": (
        "invoice number",
        "invoice no",
        "inv no",
        "document number",
        "doc no",
        "voucher number",
    ),
    "invoice_date": (
        "invoice date",
        "inv date",
        "document date",
        "doc date",
        "voucher date",
        "date",
    ),
    "invoice_value": (
        "invoice value",
        "invoice value rupees",
        "invoice value rs",
        "total invoice value",
        "total value",
        "gross amount",
    ),
    "taxable_value": (
        "taxable value",
        "taxable amount",
        "taxable value rupees",
        "taxable value rs",
    ),
    "igst": ("integrated tax", "integrated tax rupees", "igst", "igst amount"),
    "cgst": ("central tax", "central tax rupees", "cgst", "cgst amount"),
    "sgst": (
        "state/ut tax",
        "state/ut tax rupees",
        "state tax",
        "state tax rupees",
        "ut tax",
        "ut tax rupees",
        "sgst",
        "sgst amount",
    ),
    "cess": ("cess", "cess rupees", "cess amount"),
}

GSTIN_PATTERN = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.IGNORECASE)


@dataclass
class GstPurchaseRow:
    row_number: int
    source_file: str = ""
    supplier_gstin: str = ""
    supplier_name: str = ""
    invoice_number: str = ""
    invoice_date: date | None = None
    invoice_value: Decimal = Decimal("0.00")
    taxable_value: Decimal = Decimal("0.00")
    igst: Decimal = Decimal("0.00")
    cgst: Decimal = Decimal("0.00")
    sgst: Decimal = Decimal("0.00")
    cess: Decimal = Decimal("0.00")

    @property
    def tax_total(self) -> Decimal:
        return self.igst + self.cgst + self.sgst + self.cess


@dataclass
class TallyPurchaseRow:
    date: date | None
    voucher_number: str
    voucher_type_name: str
    party_ledger_name: str
    narration: str
    amount: Decimal
    guid: str = ""
    remote_id: str = ""
    voucher_key: str = ""
    master_id: str = ""
    alter_id: str = ""
    supplier_invoice_number: str = ""
    bill_references: list[str] | None = None
    supplier_gstin: str = ""
    taxable_value: Decimal = Decimal("0.00")
    igst: Decimal = Decimal("0.00")
    cgst: Decimal = Decimal("0.00")
    sgst: Decimal = Decimal("0.00")
    cess: Decimal = Decimal("0.00")
    raw_xml: str = ""

    @property
    def invoice_sources(self) -> list[str]:
        return [self.supplier_invoice_number] if self.supplier_invoice_number else []


@dataclass
class TallyMatchIndex:
    tally_rows: list[TallyPurchaseRow]
    by_invoice_text: dict[str, set[int]]
    by_invoice_suffix: dict[str, set[int]]
    by_gstin: dict[str, set[int]]
    by_amount_date: dict[tuple[Decimal, date | None], set[int]]
    by_supplier_token: dict[str, set[int]]


def parse_gst_purchase_excel(path_or_file) -> list[GstPurchaseRow]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("Excel parsing dependency missing. Please install openpyxl.") from exc

    source_file = Path(path_or_file).name if isinstance(path_or_file, (str, Path)) else getattr(path_or_file, "name", "")

    try:
        if isinstance(path_or_file, (str, Path)):
            workbook = load_workbook(path_or_file, data_only=True, read_only=True)
        else:
            path_or_file.seek(0)
            workbook = load_workbook(path_or_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("GST portal Excel file read nahi ho paayi. Valid .xlsx/.xlsm upload karein.") from exc

    sheet = workbook.active
    if sheet is None:
        raise ValueError("Excel file me active sheet nahi mili.")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file empty hai.")

    header_index, column_map = detect_gst_header_row(rows)
    if header_index is None:
        raise ValueError(
            "GST portal Excel ke headers samajh nahi aaye. Supplier GSTIN, invoice number, invoice date aur invoice value columns hone chahiye."
        )

    parsed_rows: list[GstPurchaseRow] = []
    for excel_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        parsed = parse_gst_purchase_row(row, column_map, row_number=excel_index, source_file=source_file)
        if parsed is not None:
            parsed_rows.append(parsed)

    if not parsed_rows:
        raise ValueError("GST portal Excel me valid purchase rows nahi mili.")
    return parsed_rows


def detect_gst_header_row(rows: list[tuple]) -> tuple[int | None, dict[str, int]]:
    alias_lookup = {
        normalize_header(alias): field_name
        for field_name, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    best_index = None
    best_map: dict[str, int] = {}
    best_score = 0

    for row_index, row in enumerate(rows[:30]):
        candidate_map: dict[str, int] = {}
        for cell_index, value in enumerate(row or ()):
            parent_value = rows[row_index - 1][cell_index] if row_index > 0 and cell_index < len(rows[row_index - 1]) else ""
            labels = [
                normalize_header(value),
                normalize_header(parent_value),
                normalize_header(f"{parent_value} {value}"),
            ]
            for label in labels:
                field_name = alias_lookup.get(label)
                if field_name and field_name not in candidate_map:
                    candidate_map[field_name] = cell_index
                    break

        required_score = sum(
            1
            for field_name in ("invoice_number", "invoice_date", "invoice_value")
            if field_name in candidate_map
        )
        score = required_score * 3 + len(candidate_map)
        if score > best_score:
            best_index = row_index
            best_map = candidate_map
            best_score = score

    if best_score < 8:
        return None, {}
    return best_index, best_map


def parse_gst_purchase_row(row, column_map: dict[str, int], *, row_number: int, source_file: str = "") -> GstPurchaseRow | None:
    invoice_number = read_cell(row, column_map.get("invoice_number"))
    supplier_gstin = normalize_gstin(read_cell(row, column_map.get("supplier_gstin")))
    supplier_name = read_cell(row, column_map.get("supplier_name"))
    invoice_date = parse_date_value(read_cell_raw(row, column_map.get("invoice_date")))
    invoice_value = parse_decimal_value(read_cell_raw(row, column_map.get("invoice_value")))
    taxable_value = parse_decimal_value(read_cell_raw(row, column_map.get("taxable_value")))

    if not any([invoice_number, supplier_gstin, supplier_name, invoice_date, invoice_value, taxable_value]):
        return None

    return GstPurchaseRow(
        row_number=row_number,
        source_file=source_file,
        supplier_gstin=supplier_gstin,
        supplier_name=supplier_name,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        invoice_value=invoice_value,
        taxable_value=taxable_value,
        igst=parse_decimal_value(read_cell_raw(row, column_map.get("igst"))),
        cgst=parse_decimal_value(read_cell_raw(row, column_map.get("cgst"))),
        sgst=parse_decimal_value(read_cell_raw(row, column_map.get("sgst"))),
        cess=parse_decimal_value(read_cell_raw(row, column_map.get("cess"))),
    )


def build_tally_match_index(tally_rows: list[TallyPurchaseRow]) -> TallyMatchIndex:
    by_invoice_text: dict[str, set[int]] = defaultdict(set)
    by_invoice_suffix: dict[str, set[int]] = defaultdict(set)
    by_gstin: dict[str, set[int]] = defaultdict(set)
    by_amount_date: dict[tuple[Decimal, date | None], set[int]] = defaultdict(set)
    by_supplier_token: dict[str, set[int]] = defaultdict(set)

    for index, tally_row in enumerate(tally_rows):
        for source in tally_row.invoice_sources:
            normalized_source = normalize_match_text(source)
            if len(normalized_source) >= 3:
                by_invoice_text[normalized_source].add(index)
            for token in meaningful_tokens(source):
                by_invoice_text[token].add(index)
            for suffix in invoice_digit_suffixes(source):
                by_invoice_suffix[suffix].add(index)

        tally_gstin = normalize_gstin(tally_row.supplier_gstin)
        if tally_gstin:
            by_gstin[tally_gstin].add(index)

        for amount in (tally_row.amount, tally_row.taxable_value):
            by_amount_date[(money_key(amount), tally_row.date)].add(index)

        for token in meaningful_tokens(tally_row.party_ledger_name):
            by_supplier_token[token].add(index)

    return TallyMatchIndex(
        tally_rows=tally_rows,
        by_invoice_text=dict(by_invoice_text),
        by_invoice_suffix=dict(by_invoice_suffix),
        by_gstin=dict(by_gstin),
        by_amount_date=dict(by_amount_date),
        by_supplier_token=dict(by_supplier_token),
    )


def candidate_tally_indexes(
    gst_row: GstPurchaseRow,
    match_index: TallyMatchIndex,
    unmatched_tally_indexes: set[int],
) -> set[int]:
    invoice_candidates: set[int] = set()
    suffix_candidates: set[int] = set()
    gstin_candidates: set[int] = set()
    invoice_key = normalize_match_text(gst_row.invoice_number)
    if len(invoice_key) >= 3:
        invoice_candidates.update(match_index.by_invoice_text.get(invoice_key, set()))
    for suffix in invoice_digit_suffixes(gst_row.invoice_number):
        suffix_candidates.update(match_index.by_invoice_suffix.get(suffix, set()))

    gstin = normalize_gstin(gst_row.supplier_gstin)
    if gstin:
        gstin_candidates.update(match_index.by_gstin.get(gstin, set()))

    amount_date_candidates: set[int] = set()
    if gst_row.invoice_date:
        for nearby_date in nearby_dates(gst_row.invoice_date):
            amount_date_candidates.update(match_index.by_amount_date.get((money_key(gst_row.invoice_value), nearby_date), set()))
            amount_date_candidates.update(match_index.by_amount_date.get((money_key(gst_row.taxable_value), nearby_date), set()))

    invoice_candidates &= unmatched_tally_indexes
    if invoice_candidates:
        return invoice_candidates

    suffix_candidates &= unmatched_tally_indexes
    if suffix_candidates:
        narrowed = suffix_candidates & (gstin_candidates | amount_date_candidates)
        return narrowed if narrowed else suffix_candidates

    gstin_candidates &= unmatched_tally_indexes
    if gstin_candidates:
        narrowed = gstin_candidates & amount_date_candidates
        return narrowed if narrowed else gstin_candidates

    candidates: set[int] = amount_date_candidates & unmatched_tally_indexes
    if not candidates:
        for token in meaningful_tokens(gst_row.supplier_name):
            token_candidates = match_index.by_supplier_token.get(token, set())
            if len(token_candidates) <= 80:
                candidates.update(token_candidates)
            if len(candidates) >= 80:
                break

    if not candidates and len(match_index.tally_rows) <= 250:
        return set(unmatched_tally_indexes)
    return candidates & unmatched_tally_indexes


def invoice_digit_suffixes(value: str) -> set[str]:
    digits = only_digits(value)
    if len(digits) < 2:
        return set()
    suffixes: set[str] = set()
    groups = re.findall(r"\d+", value or "")
    if groups:
        trailing_group = groups[-1]
        normalized_trailing_group = trailing_group.lstrip("0") or trailing_group
        for candidate in (trailing_group, normalized_trailing_group):
            if len(candidate) >= 2:
                suffixes.add(candidate)
    for suffix_length in (2, 3, 4, 5):
        if len(digits) >= suffix_length:
            suffixes.add(digits[-suffix_length:])
    return suffixes


def nearby_dates(value: date) -> list[date]:
    return [value + timedelta(days=offset) for offset in range(-3, 4)]


def money_key(value) -> Decimal:
    return Decimal(value or "0.00").quantize(Decimal("0.01"))


def reconcile_gst_purchases_with_tally(
    gst_rows: list[GstPurchaseRow],
    tally_vouchers,
    *,
    amount_tolerance: Decimal = Decimal("1.00"),
    tax_tolerance: Decimal = Decimal("1.00"),
    progress_callback: Callable[[int, int], None] | None = None,
) -> dict:
    tally_rows = [build_tally_purchase_row(voucher) for voucher in tally_vouchers]
    match_index = build_tally_match_index(tally_rows)
    unmatched_gst_indexes = set(range(len(gst_rows)))
    unmatched_tally_indexes = set(range(len(tally_rows)))
    results: list[dict | None] = [None] * len(gst_rows)
    score_cache: dict[tuple[int, int], dict] = {}

    for target_status in ("matched", "mismatch", "probable"):
        proposals = []
        for gst_index in sorted(unmatched_gst_indexes):
            gst_row = gst_rows[gst_index]
            for tally_index in candidate_tally_indexes(gst_row, match_index, unmatched_tally_indexes):
                cache_key = (gst_index, tally_index)
                match = score_cache.get(cache_key)
                if match is None:
                    match = score_purchase_match(
                        gst_row,
                        tally_rows[tally_index],
                        amount_tolerance=amount_tolerance,
                        tax_tolerance=tax_tolerance,
                    )
                    score_cache[cache_key] = match
                status = classify_match(match)
                if status != target_status:
                    continue
                proposals.append(
                    (
                        match["match_rank"],
                        match["score"],
                        match["identity_weight"],
                        -gst_index,
                        -tally_index,
                        gst_index,
                        tally_index,
                        match,
                        status,
                    )
                )

        proposals.sort(reverse=True)
        for *_ranking, gst_index, tally_index, match, status in proposals:
            if gst_index not in unmatched_gst_indexes or tally_index not in unmatched_tally_indexes:
                continue
            results[gst_index] = build_match_result(status, gst_rows[gst_index], match)
            unmatched_gst_indexes.remove(gst_index)
            unmatched_tally_indexes.remove(tally_index)

    for gst_index in sorted(unmatched_gst_indexes):
        results[gst_index] = {
            "status": "missing",
            "gst": gst_rows[gst_index],
            "tally": None,
            "score": 0,
            "reasons": ["No reliable Tally match found"],
        }

    if progress_callback is not None:
        for processed_count in range(1, len(gst_rows) + 1):
            progress_callback(processed_count, len(gst_rows))

    tally_only = [
        tally_row
        for index, tally_row in enumerate(tally_rows)
        if index in unmatched_tally_indexes
    ]
    final_results = [result for result in results if result is not None]
    summary = {
        "gst_count": len(gst_rows),
        "tally_count": len(tally_rows),
        "matched_count": sum(1 for result in final_results if result["status"] == "matched"),
        "probable_count": sum(1 for result in final_results if result["status"] == "probable"),
        "mismatch_count": sum(1 for result in final_results if result["status"] == "mismatch"),
        "missing_count": sum(1 for result in final_results if result["status"] == "missing"),
        "tally_only_count": len(tally_only),
    }
    return {"results": final_results, "tally_only": tally_only, "summary": summary}


def build_match_result(status: str, gst_row: GstPurchaseRow, match: dict) -> dict:
    return {
        "status": status,
        "gst": gst_row,
        "tally": match["tally"],
        "score": match["score"],
        "reasons": match["reasons"],
    }


def score_purchase_match(
    gst_row: GstPurchaseRow,
    tally_row: TallyPurchaseRow,
    *,
    amount_tolerance: Decimal,
    tax_tolerance: Decimal,
) -> dict:
    score = 0
    identity_weight = 0
    reasons: list[str] = []
    gstin_match = False
    tally_gstin = normalize_gstin(tally_row.supplier_gstin)
    gst_gstin = normalize_gstin(gst_row.supplier_gstin)

    if gst_gstin and tally_gstin and gst_gstin == tally_gstin:
        score += 28
        identity_weight += 5
        gstin_match = True
        reasons.append("GSTIN matched")
    elif gst_gstin and not tally_gstin:
        reasons.append("Tally GSTIN missing")
    elif gst_gstin and tally_gstin:
        score -= 25
        reasons.append(f"GSTIN mismatch: Tally {tally_gstin}")

    full_invoice_match = invoice_full_match(gst_row.invoice_number, tally_row.invoice_sources)
    suffix_invoice_match = invoice_suffix_match(gst_row.invoice_number, tally_row.invoice_sources)
    if full_invoice_match:
        score += 70
        identity_weight += 3
        reasons.append("Invoice number matched")
    elif suffix_invoice_match:
        score += 42
        identity_weight += 2
        reasons.append("Invoice suffix matched")

    supplier_score = supplier_name_score(gst_row.supplier_name, tally_row.party_ledger_name)
    if supplier_score >= 2:
        score += 22
        identity_weight += 2
        reasons.append("Supplier name similar")
    elif supplier_score == 1:
        score += 12
        identity_weight += 1
        reasons.append("Supplier name partially matched")

    amount_checks = compare_amounts(gst_row, tally_row, amount_tolerance=amount_tolerance, tax_tolerance=tax_tolerance)
    amount_lookup = {label: ok for label, ok, _diff, _tolerance in amount_checks}
    for label, ok, diff, tolerance in amount_checks:
        if ok:
            if is_material_amount_check(label, gst_row, tally_row):
                score += 10
                reasons.append(f"{label} matched")
        else:
            score -= 18
            reasons.append(f"{label} mismatch Rs {diff:.2f} (limit Rs {tolerance:.2f})")

    date_close = False
    date_mismatch = False
    if gst_row.invoice_date and tally_row.date:
        day_difference = abs((gst_row.invoice_date - tally_row.date).days)
        if day_difference == 0:
            score += 12
            date_close = True
            reasons.append("Invoice date exact")
        elif day_difference <= 3:
            score += 8
            date_close = True
            reasons.append(f"Invoice date close by {day_difference} days")
        else:
            score -= 8
            date_mismatch = True
            reasons.append(f"Invoice date mismatch by {day_difference} days")

    invoice_value_ok = amount_lookup.get("Invoice value", False)
    taxable_value_ok = amount_lookup.get("Taxable value", False)
    amount_tax_ok = all(
        ok
        for label, ok, _diff, _tolerance in amount_checks
        if is_material_amount_check(label, gst_row, tally_row)
    )
    same_party = gstin_match or supplier_score > 0
    has_date_pair = bool(gst_row.invoice_date and tally_row.date)
    invoice_evidence = full_invoice_match or suffix_invoice_match
    suffix_supported = suffix_invoice_match and same_party and amount_tax_ok and not date_mismatch and (
        date_close if has_date_pair else True
    )
    suffix_tax_review_supported = suffix_invoice_match and same_party and invoice_value_ok and not date_mismatch and (
        date_close if has_date_pair else True
    )
    identity_evidence = (
        (full_invoice_match and (same_party or amount_tax_ok))
        or suffix_supported
        or suffix_tax_review_supported
    )
    if suffix_invoice_match and not suffix_supported and not suffix_tax_review_supported:
        reasons.append("Invoice suffix ignored because amount/date support missing")

    match_rank = 0
    if full_invoice_match:
        match_rank = 4
    elif suffix_supported or suffix_tax_review_supported:
        match_rank = 3

    if not reasons:
        reasons.append("Weak match")

    return {
        "score": score,
        "identity_weight": identity_weight,
        "match_rank": match_rank,
        "reasons": reasons,
        "tally": tally_row,
        "gstin_match": gstin_match,
        "full_invoice_match": full_invoice_match,
        "suffix_invoice_match": suffix_invoice_match,
        "invoice_evidence": invoice_evidence,
        "suffix_supported": suffix_supported,
        "suffix_tax_review_supported": suffix_tax_review_supported,
        "supplier_score": supplier_score,
        "same_party": same_party,
        "amount_tax_ok": amount_tax_ok,
        "date_close": date_close,
        "date_mismatch": date_mismatch,
        "identity_evidence": identity_evidence,
    }


def classify_match(match: dict | None) -> str:
    if not match or match["match_rank"] <= 0:
        return "missing"
    if not match["identity_evidence"]:
        return "missing"

    if match["full_invoice_match"]:
        if not match["amount_tax_ok"] or match["date_mismatch"]:
            return "mismatch"
        return "matched"

    if match["suffix_supported"]:
        return "probable"

    if match.get("suffix_tax_review_supported"):
        return "probable"

    return "missing"


def compare_amounts(
    gst_row: GstPurchaseRow,
    tally_row: TallyPurchaseRow,
    *,
    amount_tolerance: Decimal,
    tax_tolerance: Decimal,
) -> list[tuple[str, bool, Decimal, Decimal]]:
    checks = [
        ("Invoice value", gst_row.invoice_value, tally_row.amount, amount_tolerance),
        ("Taxable value", gst_row.taxable_value, tally_row.taxable_value, tax_tolerance),
        ("IGST", gst_row.igst, tally_row.igst, tax_tolerance),
        ("CGST", gst_row.cgst, tally_row.cgst, tax_tolerance),
        ("SGST", gst_row.sgst, tally_row.sgst, tax_tolerance),
        ("Cess", gst_row.cess, tally_row.cess, tax_tolerance),
    ]
    results = []
    for label, gst_amount, tally_amount, tolerance in checks:
        diff = abs(Decimal(gst_amount or "0.00") - Decimal(tally_amount or "0.00")).quantize(Decimal("0.01"))
        results.append((label, diff <= tolerance, diff, tolerance))
    return results


def is_material_amount_check(label: str, gst_row: GstPurchaseRow, tally_row: TallyPurchaseRow) -> bool:
    if label in {"Invoice value", "Taxable value"}:
        return True

    field_name = {
        "IGST": "igst",
        "CGST": "cgst",
        "SGST": "sgst",
        "Cess": "cess",
    }.get(label)
    if not field_name:
        return True
    return abs(Decimal(getattr(gst_row, field_name) or "0.00")) > Decimal("0.00") or abs(
        Decimal(getattr(tally_row, field_name) or "0.00")
    ) > Decimal("0.00")


def build_tally_purchase_row(voucher) -> TallyPurchaseRow:
    return TallyPurchaseRow(
        date=parse_date_value(getattr(voucher, "date", "")),
        voucher_number=getattr(voucher, "voucher_number", "") or "",
        supplier_invoice_number=getattr(voucher, "supplier_invoice_number", "") or "",
        voucher_type_name=getattr(voucher, "voucher_type_name", "") or "",
        party_ledger_name=getattr(voucher, "party_ledger_name", "") or "",
        narration=getattr(voucher, "narration", "") or "",
        amount=abs(parse_decimal_value(getattr(voucher, "amount", ""))),
        guid=getattr(voucher, "guid", "") or "",
        remote_id=getattr(voucher, "remote_id", "") or "",
        voucher_key=getattr(voucher, "voucher_key", "") or "",
        master_id=getattr(voucher, "master_id", "") or "",
        alter_id=getattr(voucher, "alter_id", "") or "",
        bill_references=list(getattr(voucher, "bill_references", []) or []),
        supplier_gstin=normalize_gstin(getattr(voucher, "supplier_gstin", "") or ""),
        taxable_value=abs(parse_decimal_value(getattr(voucher, "taxable_value", ""))),
        igst=abs(parse_decimal_value(getattr(voucher, "igst", ""))),
        cgst=abs(parse_decimal_value(getattr(voucher, "cgst", ""))),
        sgst=abs(parse_decimal_value(getattr(voucher, "sgst", ""))),
        cess=abs(parse_decimal_value(getattr(voucher, "cess", ""))),
        raw_xml=getattr(voucher, "raw_xml", "") or "",
    )


def invoice_full_match(gst_invoice: str, sources: list[str]) -> bool:
    invoice = normalize_match_text(gst_invoice)
    if not invoice:
        return False
    invoice_digits = only_digits(invoice)
    if len(invoice) < 3:
        return any(
            invoice == normalize_match_text(source)
            or (len(invoice_digits) >= 2 and invoice_digits == only_digits(source))
            for source in sources
        )
    for source in sources:
        source_text = normalize_match_text(source)
        source_digits = only_digits(source)
        if invoice == source_text:
            return True
        if invoice.isdigit() and invoice_digits and invoice_digits == source_digits:
            return True
        if not invoice.isdigit() and invoice in source_text:
            return True
    return False


def invoice_suffix_match(gst_invoice: str, sources: list[str]) -> bool:
    suffixes = invoice_digit_suffixes(gst_invoice)
    if not suffixes:
        return False
    for source in sources:
        source_digits = only_digits(source)
        if not source_digits:
            continue
        for suffix in suffixes:
            if suffix and source_digits.endswith(suffix):
                return True
    return False


def supplier_name_score(gst_supplier_name: str, tally_party_name: str) -> int:
    gst_tokens = meaningful_tokens(gst_supplier_name)
    tally_tokens = meaningful_tokens(tally_party_name)
    if not gst_tokens or not tally_tokens:
        return 0
    common = gst_tokens & tally_tokens
    if len(common) >= 2:
        return 2
    if common:
        return 1
    return 0


def meaningful_tokens(value: str) -> set[str]:
    stop_words = {"and", "the", "pvt", "ltd", "limited", "private", "co", "company", "m/s", "ms", "shri", "sri"}
    return {
        token
        for token in re.findall(r"[a-z0-9]+", (value or "").lower())
        if len(token) >= 3 and token not in stop_words
    }


def normalize_gstin(value: str) -> str:
    match = GSTIN_PATTERN.search((value or "").upper())
    return match.group(0).upper() if match else (value or "").strip().upper()


def normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("(₹)", " rupees ").replace("(rs.)", " rupees ").replace("(rs)", " rupees ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_match_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def only_digits(value: str) -> str:
    return re.sub(r"\D+", "", value or "")


def read_cell(row, index: int | None) -> str:
    raw_value = read_cell_raw(row, index)
    if raw_value is None:
        return ""
    return str(raw_value).strip()


def read_cell_raw(row, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_decimal_value(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).replace(",", "").strip()
    if not text:
        return Decimal("0.00")
    is_negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return Decimal("0.00")
    try:
        amount = Decimal(match.group(0)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")
    return -abs(amount) if is_negative else amount


def parse_date_value(value) -> date | None:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).date()
        except Exception:
            return None

    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
