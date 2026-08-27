from __future__ import annotations

import random
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_CEILING, ROUND_HALF_UP


def extract_gst_rate_from_stock_item_name(stock_item_name: str) -> Decimal:
    stock_name = (stock_item_name or "").strip()
    if not stock_name:
        return Decimal("0.00")

    match = re.search(r"(\d+(?:\.\d+)?)\s*%", stock_name)
    if not match:
        match = re.search(r"(\d+(?:\.\d+)?)", stock_name)
    if not match:
        return Decimal("0.00")

    return Decimal(match.group(1)).quantize(Decimal("0.01"))


def split_gst_rate(total_gst_rate: Decimal, *, has_cgst: bool, has_sgst: bool) -> tuple[Decimal, Decimal]:
    total_rate = Decimal(total_gst_rate or "0.00").quantize(Decimal("0.01"))

    if total_rate <= Decimal("0.00"):
        return Decimal("0.00"), Decimal("0.00")
    if has_cgst and has_sgst:
        half_rate = (total_rate / Decimal("2.00")).quantize(Decimal("0.01"))
        return half_rate, half_rate
    if has_cgst:
        return total_rate, Decimal("0.00")
    if has_sgst:
        return Decimal("0.00"), total_rate
    return Decimal("0.00"), Decimal("0.00")


def pick_spread_date(
    index: int,
    *,
    from_date: date,
    to_date: date,
    number_of_bills: int,
) -> date:
    date_range = (to_date - from_date).days
    if date_range == 0:
        return from_date
    step = date_range / max(number_of_bills - 1, 1)
    offset = round(step * index)
    return from_date + timedelta(days=min(offset, date_range))


def calculate_tax_components(
    base_amount: Decimal,
    *,
    cgst_rate: Decimal = Decimal("0.00"),
    sgst_rate: Decimal = Decimal("0.00"),
    has_round_off: bool,
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    quant = Decimal("0.01")
    cgst_amount = (base_amount * (cgst_rate / Decimal("100.00"))).quantize(quant)
    sgst_amount = (base_amount * (sgst_rate / Decimal("100.00"))).quantize(quant)
    invoice_total = (base_amount + cgst_amount + sgst_amount).quantize(quant)
    round_off_amount = Decimal("0.00")
    if has_round_off:
        rounded_total = invoice_total.quantize(Decimal("1"), rounding=ROUND_CEILING).quantize(quant)
        round_off_amount = (rounded_total - invoice_total).quantize(quant)
    gross_amount = (invoice_total + round_off_amount).quantize(quant)
    return cgst_amount, sgst_amount, round_off_amount, gross_amount


def gross_amount_from_taxable(
    taxable_amount: Decimal,
    *,
    cgst_rate: Decimal = Decimal("0.00"),
    sgst_rate: Decimal = Decimal("0.00"),
    has_round_off: bool,
) -> Decimal:
    return calculate_tax_components(
        taxable_amount,
        cgst_rate=cgst_rate,
        sgst_rate=sgst_rate,
        has_round_off=has_round_off,
    )[3]


def split_random_sale_taxable_amounts(
    total_taxable_amount: Decimal,
    *,
    number_of_bills: int,
    min_amount: Decimal,
    max_amount: Decimal,
    cgst_rate: Decimal = Decimal("0.00"),
    sgst_rate: Decimal = Decimal("0.00"),
    has_round_off: bool,
    rng: random.Random | None = None,
    max_attempts: int = 25,
) -> list[Decimal]:
    quant = Decimal("0.01")
    rng = rng or random.Random()

    min_cents = int((min_amount * 100).to_integral_value())
    max_cents = int((max_amount * 100).to_integral_value())
    total_cents = int((total_taxable_amount * 100).to_integral_value())
    min_total_cents = min_cents * number_of_bills
    max_total_cents = max_cents * number_of_bills
    if not (min_total_cents <= total_cents <= max_total_cents):
        return [(total_taxable_amount / number_of_bills).quantize(quant) for _ in range(number_of_bills)]

    if number_of_bills <= 1:
        return [total_taxable_amount.quantize(quant)]

    extra_cents = total_cents - min_total_cents
    capacity_per_bill = max_cents - min_cents
    if capacity_per_bill <= 0:
        return [Decimal(min_cents) / Decimal("100") for _ in range(number_of_bills)]

    base_extra_cents = extra_cents // number_of_bills
    remainder_cents = extra_cents % number_of_bills
    amounts_cents = [min_cents + base_extra_cents for _ in range(number_of_bills)]
    if remainder_cents:
        bonus_indexes = rng.sample(range(number_of_bills), remainder_cents)
        for index in bonus_indexes:
            amounts_cents[index] += 1

    jitter_budget = max(number_of_bills * 12, extra_cents // 2)
    transfer_cap = max(1, capacity_per_bill // 6)
    for _ in range(min(max_attempts * number_of_bills, jitter_budget)):
        donor_candidates = [idx for idx, value in enumerate(amounts_cents) if value > min_cents]
        receiver_candidates = [idx for idx, value in enumerate(amounts_cents) if value < max_cents]
        if not donor_candidates or not receiver_candidates:
            break

        donor_index = rng.choice(donor_candidates)
        receiver_index = rng.choice(receiver_candidates)
        if donor_index == receiver_index:
            continue

        donor_available = amounts_cents[donor_index] - min_cents
        receiver_space = max_cents - amounts_cents[receiver_index]
        if donor_available <= 0 or receiver_space <= 0:
            continue

        max_transfer = min(donor_available, receiver_space, transfer_cap)
        if max_transfer <= 0:
            continue

        transfer = max(1, int(round(max_transfer * (rng.random() ** 0.35))))
        transfer = min(transfer, max_transfer)

        amounts_cents[donor_index] -= transfer
        amounts_cents[receiver_index] += transfer

    rng.shuffle(amounts_cents)
    return [(Decimal(c) / Decimal("100")).quantize(quant) for c in amounts_cents]


def spread_random_sale_dates(
    *,
    from_date: date,
    to_date: date,
    number_of_bills: int,
    rng: random.Random | None = None,
) -> list[date]:
    rng = rng or random.Random()
    date_range = (to_date - from_date).days
    if number_of_bills <= 1 or date_range <= 0:
        return [from_date for _ in range(number_of_bills)]

    offsets = [0, date_range]
    interior_slots = number_of_bills - 2
    if interior_slots > 0:
        possible_offsets = list(range(1, date_range))
        if len(possible_offsets) >= interior_slots:
            offsets.extend(rng.sample(possible_offsets, interior_slots))
        else:
            offsets.extend(rng.randint(1, date_range - 1) for _ in range(interior_slots))

    offsets.sort()
    return [from_date + timedelta(days=offset) for offset in offsets]


def build_random_sale_preview_entries(
    *,
    party_name: str,
    stock_item_name: str,
    sale_ledger: str,
    voucher_type: str,
    cash_bank_ledger: str,
    cgst_ledger: str,
    sgst_ledger: str,
    round_off_ledger: str,
    narration_prefix: str,
    from_date: date,
    to_date: date,
    number_of_bills: int,
    taxable_amount: Decimal,
    min_amount: Decimal,
    max_amount: Decimal,
    cgst_rate: Decimal | None = None,
    sgst_rate: Decimal | None = None,
    rng: random.Random | None = None,
) -> list[dict]:
    party_ledger = party_name.strip() or cash_bank_ledger.strip()
    if not party_ledger:
        raise ValueError("Party ya Cash / Bank ledger me se ek required hai.")

    if cgst_rate is None or sgst_rate is None:
        total_gst_rate = extract_gst_rate_from_stock_item_name(stock_item_name)
        cgst_rate, sgst_rate = split_gst_rate(
            total_gst_rate,
            has_cgst=bool(cgst_ledger.strip()),
            has_sgst=bool(sgst_ledger.strip()),
        )
    else:
        cgst_rate = Decimal(cgst_rate).quantize(Decimal("0.01")) if cgst_ledger.strip() else Decimal("0.00")
        sgst_rate = Decimal(sgst_rate).quantize(Decimal("0.01")) if sgst_ledger.strip() else Decimal("0.00")

    taxable_amounts = split_random_sale_taxable_amounts(
        taxable_amount,
        number_of_bills=number_of_bills,
        min_amount=min_amount,
        max_amount=max_amount,
        cgst_rate=cgst_rate,
        sgst_rate=sgst_rate,
        has_round_off=bool(round_off_ledger.strip()),
        rng=rng,
    )

    entry_dates = spread_random_sale_dates(
        from_date=from_date,
        to_date=to_date,
        number_of_bills=number_of_bills,
        rng=rng,
    )

    preview_entries: list[dict] = []
    clean_narration_prefix = narration_prefix.strip()
    for index, entry_taxable_amount in enumerate(taxable_amounts, start=1):
        cgst_amount, sgst_amount, round_off_amount, gross_amount = calculate_tax_components(
            entry_taxable_amount,
            cgst_rate=cgst_rate,
            sgst_rate=sgst_rate,
            has_round_off=bool(round_off_ledger.strip()),
        )
        preview_entries.append(
            {
                "bill_number": index,
                "voucher_date": entry_dates[index - 1],
                "party_name": party_ledger,
                "party_ledger": party_ledger,
                "stock_item_name": stock_item_name,
                "sale_ledger": sale_ledger,
                "voucher_type": voucher_type,
                "cash_bank_ledger": cash_bank_ledger,
                "cgst_ledger": cgst_ledger,
                "sgst_ledger": sgst_ledger,
                "cgst_rate": cgst_rate,
                "sgst_rate": sgst_rate,
                "round_off_ledger": round_off_ledger,
                "taxable_amount": entry_taxable_amount,
                "cgst_amount": cgst_amount,
                "sgst_amount": sgst_amount,
                "round_off_amount": round_off_amount,
                "amount": gross_amount,
                "narration": (
                    f"{clean_narration_prefix} {index} - {party_ledger} - {stock_item_name}"
                    if clean_narration_prefix
                    else ""
                ),
            }
        )

    return preview_entries


def parse_fixed_sale_excel_rows(path_or_file) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore[reportMissingModuleSource]
        from openpyxl.utils.datetime import from_excel  # type: ignore[reportMissingModuleSource]
    except Exception as exc:
        raise ValueError("Excel parsing dependency missing. Please install openpyxl.") from exc

    try:
        if hasattr(path_or_file, "seek"):
            path_or_file.seek(0)
        workbook = load_workbook(path_or_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("Excel file read nahi ho paayi. Valid .xlsx file select karein.") from exc

    sheet = workbook.active
    if sheet is None:
        raise ValueError("Excel file me active sheet nahi mili.")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file empty hai. Date aur Amount rows add karein.")

    header = rows[0] or ()
    header_map: dict[str, int] = {}
    for idx, value in enumerate(header):
        key = str(value).strip().lower() if value is not None else ""
        if key:
            header_map[key] = idx

    def find_header_index(*names: str) -> int | None:
        for name in names:
            if name in header_map:
                return header_map[name]
        return None

    date_idx = find_header_index("date", "voucher date", "transaction date", "bill date")
    amount_idx = find_header_index(
        "amount",
        "debit",
        "debit amount",
        "credit",
        "credit amount",
        "total",
        "gross amount",
    )
    particulars_idx = find_header_index("particulars", "particular", "party", "party name")
    narration_idx = find_header_index("narration", "description", "remarks", "remark")

    has_named_header = date_idx is not None and amount_idx is not None
    date_idx = 0 if date_idx is None else date_idx
    amount_idx = 1 if amount_idx is None else amount_idx
    start_row_idx = 1 if has_named_header else 0

    parsed_rows: list[dict] = []
    for excel_row_number, row in enumerate(rows[start_row_idx:], start=start_row_idx + 1):
        if not row:
            continue

        date_cell = row[date_idx] if len(row) > date_idx else None
        amount_cell = row[amount_idx] if len(row) > amount_idx else None

        if date_cell in (None, "") and amount_cell in (None, ""):
            continue
        if date_cell in (None, ""):
            raise ValueError(f"Row {excel_row_number}: Date missing hai.")
        if amount_cell in (None, ""):
            raise ValueError(f"Row {excel_row_number}: Amount missing hai.")

        parsed_date = parse_fixed_sale_date(date_cell, excel_row_number, from_excel)
        parsed_amount = parse_fixed_sale_amount(amount_cell, excel_row_number)

        source_particulars = ""
        if particulars_idx is not None and len(row) > particulars_idx and row[particulars_idx] is not None:
            source_particulars = str(row[particulars_idx]).strip()

        source_narration = ""
        if narration_idx is not None and len(row) > narration_idx and row[narration_idx] is not None:
            source_narration = str(row[narration_idx]).strip()

        parsed_rows.append(
            {
                "voucher_date": parsed_date,
                "amount": parsed_amount,
                "source_particulars": source_particulars,
                "source_narration": source_narration,
            }
        )

    if not parsed_rows:
        raise ValueError("Excel file me valid rows nahi mili. Date aur Amount columns bharein.")

    return parsed_rows


def parse_fixed_sale_date(date_cell, excel_row_number: int, from_excel) -> date:
    if isinstance(date_cell, datetime):
        return date_cell.date()
    if isinstance(date_cell, date):
        return date_cell
    if isinstance(date_cell, (int, float)):
        try:
            parsed_excel_date = from_excel(date_cell)
        except Exception as exc:
            raise ValueError(f"Row {excel_row_number}: Date format invalid hai. Excel date sahi nahi hai.") from exc
        if isinstance(parsed_excel_date, datetime):
            return parsed_excel_date.date()
        if isinstance(parsed_excel_date, date):
            return parsed_excel_date

    raw_date = str(date_cell).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_date, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Row {excel_row_number}: Date format invalid hai. Use YYYY-MM-DD ya DD-MM-YYYY.")


def parse_fixed_sale_amount(amount_cell, excel_row_number: int) -> Decimal:
    try:
        parsed_amount = Decimal(str(amount_cell).replace(",", "").strip()).quantize(Decimal("0.01"))
    except Exception as exc:
        raise ValueError(f"Row {excel_row_number}: Amount number nahi hai.") from exc

    if parsed_amount <= Decimal("0.00"):
        raise ValueError(f"Row {excel_row_number}: Amount zero se bada hona chahiye.")
    return parsed_amount


def build_fixed_sale_narration(
    *,
    narration_prefix: str,
    bill_number: int,
    party_name: str,
    stock_item_name: str,
    source_particulars: str = "",
    source_narration: str = "",
) -> str:
    narration_parts: list[str] = []
    seen_parts: set[str] = set()

    for raw_part in (narration_prefix, source_narration):
        cleaned_part = str(raw_part or "").strip()
        if not cleaned_part:
            continue

        normalized_part = " ".join(cleaned_part.split()).casefold()
        if normalized_part in seen_parts:
            continue

        seen_parts.add(normalized_part)
        narration_parts.append(cleaned_part)

    return " | ".join(narration_parts)


def derive_invoice_components_from_total(
    target_total: Decimal,
    *,
    cgst_rate: Decimal,
    sgst_rate: Decimal,
    has_round_off: bool,
) -> tuple[Decimal, Decimal, Decimal, Decimal, Decimal]:
    quant = Decimal("0.01")
    total_q = Decimal(target_total).quantize(quant)
    total_tax_percentage = Decimal(cgst_rate or "0.00") + Decimal(sgst_rate or "0.00")

    if total_tax_percentage <= Decimal("0.00"):
        return total_q, Decimal("0.00"), Decimal("0.00"), Decimal("0.00"), total_q

    divisor = Decimal("1.00") + (total_tax_percentage / Decimal("100.00"))
    taxable_amount = (total_q / divisor).quantize(quant, rounding=ROUND_HALF_UP)
    cgst_amount = (taxable_amount * (Decimal(cgst_rate) / Decimal("100.00"))).quantize(
        quant,
        rounding=ROUND_HALF_UP,
    )
    sgst_amount = (taxable_amount * (Decimal(sgst_rate) / Decimal("100.00"))).quantize(
        quant,
        rounding=ROUND_HALF_UP,
    )
    subtotal = (taxable_amount + cgst_amount + sgst_amount).quantize(quant)

    if has_round_off:
        round_off_amount = (total_q - subtotal).quantize(quant)
        gross_amount = total_q
    else:
        round_off_amount = Decimal("0.00")
        gross_amount = subtotal

    return taxable_amount, cgst_amount, sgst_amount, round_off_amount, gross_amount


def build_fixed_sale_preview_entries(
    excel_rows: list[dict],
    *,
    party_name: str,
    stock_item_name: str,
    sale_ledger: str,
    voucher_type: str,
    cash_bank_ledger: str,
    cgst_ledger: str,
    sgst_ledger: str,
    round_off_ledger: str,
    narration_prefix: str,
    cgst_rate: Decimal | None = None,
    sgst_rate: Decimal | None = None,
    item_rate: Decimal | None = None,
    billed_quantity: Decimal | None = None,
) -> list[dict]:
    party_ledger = party_name.strip() or cash_bank_ledger.strip()
    if not party_ledger:
        raise ValueError("Party ya Cash / Bank ledger me se ek required hai.")

    if cgst_rate is None or sgst_rate is None:
        stock_gst_rate = extract_gst_rate_from_stock_item_name(stock_item_name)
        cgst_rate, sgst_rate = split_gst_rate(
            stock_gst_rate,
            has_cgst=bool(cgst_ledger.strip()),
            has_sgst=bool(sgst_ledger.strip()),
        )
    else:
        cgst_rate = Decimal(cgst_rate).quantize(Decimal("0.01")) if cgst_ledger.strip() else Decimal("0.00")
        sgst_rate = Decimal(sgst_rate).quantize(Decimal("0.01")) if sgst_ledger.strip() else Decimal("0.00")

    preview_entries: list[dict] = []
    for index, row in enumerate(excel_rows, start=1):
        taxable_amount, cgst_amount, sgst_amount, round_off_amount, gross_amount = derive_invoice_components_from_total(
            row["amount"],
            cgst_rate=cgst_rate,
            sgst_rate=sgst_rate,
            has_round_off=bool(round_off_ledger.strip()),
        )
        entry_quantity = billed_quantity
        if item_rate is not None and entry_quantity is None:
            entry_quantity = (taxable_amount / item_rate).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
            if entry_quantity <= Decimal("0.000"):
                entry_quantity = Decimal("0.001")

        preview_entries.append(
            {
                "bill_number": index,
                "voucher_date": row["voucher_date"],
                "party_name": party_ledger,
                "party_ledger": party_ledger,
                "stock_item_name": stock_item_name,
                "sale_ledger": sale_ledger,
                "voucher_type": voucher_type,
                "cash_bank_ledger": cash_bank_ledger,
                "cgst_ledger": cgst_ledger,
                "sgst_ledger": sgst_ledger,
                "cgst_rate": cgst_rate,
                "sgst_rate": sgst_rate,
                "round_off_ledger": round_off_ledger,
                "taxable_amount": taxable_amount,
                "cgst_amount": cgst_amount,
                "sgst_amount": sgst_amount,
                "round_off_amount": round_off_amount,
                "amount": gross_amount,
                "source_particulars": row.get("source_particulars", ""),
                "source_narration": row.get("source_narration", ""),
                "item_rate": item_rate,
                "billed_quantity": entry_quantity,
                "narration": build_fixed_sale_narration(
                    narration_prefix=narration_prefix,
                    bill_number=index,
                    party_name=party_ledger,
                    stock_item_name=stock_item_name,
                    source_particulars=row.get("source_particulars", ""),
                    source_narration=row.get("source_narration", ""),
                ),
            }
        )

    return preview_entries
